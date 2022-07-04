import sys
import os
import getopt

from intuitlib.client import AuthClient
from intuitlib.enums import Scopes
from quickbooks import QuickBooks
from quickbooks.objects.customer import Customer
from quickbooks.objects.invoice import Invoice
from quickbooks.objects.department import Department
from quickbooks.objects.item import Item
from quickbooks.objects.term import Term
from quickbooks.objects.classitem import ClassItem
from quickbooks.objects.creditmemo import CreditMemo

from openpyxl import Workbook, load_workbook

from settings import *


actions = {
    'invoices': ["-i", "-o"],
    'locations': [],
}

all_arguments = []
for name, arguments in actions.items():
    all_arguments.extend(arguments)


def main(argv):
    global _arguments
    _arguments = {}
    try:
        global _action
        _action = argv[0].lower()
        if _action not in actions.keys():
            print("ERROR: unknown action %s" % _action)
            usage()
            sys.exit(2)
        argv = argv[1:]
        opts, args = getopt.getopt(argv, ':'.join(
            all_arguments).replace('-', '') + ':')
        if opts:
            for opt, arg in opts:
                if opt in actions[_action]:
                    _arguments[opt] = arg
    except getopt.GetoptError:
        usage()
        sys.exit(2)

    # check if all arguements are set for the action
    if set(actions[_action]) != set(_arguments.keys()):
        print("ERROR: missing arguments for command %s" % _action)
        usage()
        sys.exit(2)


def usage():
    """Print command usage"""
    print()
    print("""******** RSTRT Python Quickbooks Script ********
usage:

Set invoice parameters from input file
python main.py import -i input_file -o output_file

Set location in invoices
python main.py locations

""")





class PythonQuickBooks(object):
    """
    PythonQuickBooks

    Main class that acts as in interface for the quickbooks library
    """

    def __init__(self):
        """
        Init function

        Create a client for the API and load all the clients
        to speed up the retrival process.

        TODO now it loads only 1000 clients, extend to all
        """

        super(PythonQuickBooks, self).__init__()
        self.client = self._create_client()
        self.customers = self._load_customers()
        self.invoices = self._load_invoices()
        self.creditnotes = self._load_creditnotes()
        self.trovati = 0

    def _load_customers(self):
        customers = Customer.where(
            "Active=True",
            order_by='DisplayName',
            max_results=1000,
            start_position=1,
            qb=self.client)

        customers += Customer.where(
            "Active=True",
            order_by='DisplayName',
            max_results=1000,
            start_position=1000,
            qb=self.client)

        return customers

    def _load_invoices(self):
        invoices = Invoice.where(
            max_results=1000,
            start_position=1,
            order_by="TxnDate DESC",
            qb=self.client)

        invoices += Invoice.where(
            max_results=1000,
            start_position=1000,
            order_by="TxnDate DESC",
            qb=self.client)

        return invoices

    def _load_creditnotes(self):
        creditnotes = CreditMemo.where(
            max_results=1000,
            start_position=1,
            order_by="TxnDate DESC",
            qb=self.client)

        creditnotes += CreditMemo.where(
            max_results=1000,
            start_position=1000,
            order_by="TxnDate DESC",
            qb=self.client)

        return creditnotes

    def _create_client(self):
        auth_client = AuthClient(
            client_id=CLIENT_ID,
            client_secret=CLIENT_SECRET,
            environment=ENVIRONMENT,
            redirect_uri='https://developer.intuit.com/v2/OAuth2Playground/RedirectUrl',
        )

        # url = auth_client.get_authorization_url([Scopes.ACCOUNTING])
        # res = auth_client.get(url=url)
        # print(help(res))
        # sys.exit()
        # print(f"access_token: {auth_client.access_token}")
        # print(f"refresh_token: {auth_client.refresh_token}")
        # print(f"realm_id: {auth_client.realm_id}")



        # print(help(auth_client))

        # auth_code = request.GET.get('code', None)
        # realm_id = request.GET.get('realmId', None)
        # request.session['realm_id'] = realm_id

        # # Refresh token endpoint
        # auth_client.refresh(refresh_token=REFRESH_TOKEN)

        client = QuickBooks(
            auth_client=auth_client,
            refresh_token=REFRESH_TOKEN,
            company_id=COMPANY_ID,
            minorversion=54
        )

        # return None
        return client

    def _get_location(self, sales_name):
        try:
            location = Department.filter(Name=sales_name, qb=self.client)[0]
            return location
        except:
            return None

    def _set_location_in_invoices(self):
        """
        Change Location value in invoice to Sales
        contained in client's suffix field

        TODO: make it scriptable to run into a cron job
        """
        print(f"set_location_in_invoices started...")
        for i in self.invoices:
            if i.DepartmentRef is None:
                c = Customer.get(i.CustomerRef.value, qb=self.client)
                if c.Suffix != "":
                    location = self._get_location(sales_name=c.Suffix)
                    if location is not None:
                        # print(f"INVOICE {i.DocNumber} ASSIGNED TO {c.Suffix} LOCATION ID {location.Id}")
                        i.DepartmentRef = location.to_ref()
                        try:
                            i.save(qb=self.client)
                        except Exception as e:
                            print(f"[ERROR][I:{i.DocNumber}][C:{c}][A:{i.TotalAmt}] - {e}")
                            # print(f"FATTURA:{i.DocNumber} Località: {i.DepartmentRef}")
                    else:
                        print(f"[ERROR] Location IS NONE FOR {c.Suffix} OF {c} - {c.Active}")
                else:
                    print(f"[ERROR] c.Suffix IS NONE FOR {c}")
        print(f"set_location_in_invoices ended.")

    def _set_location_in_creditnotes(self):
        """
        Change Location value in creditnotes to Sales
        contained in client's suffix field

        TODO: make it scriptable to run into a cron job
        """
        print(f"set_location_in_creditnotes started...")
        for cn in self.creditnotes:
            if cn.DepartmentRef is None:
                c = Customer.get(cn.CustomerRef.value, qb=self.client)
                if c.Suffix != "":
                    location = self._get_location(sales_name=c.Suffix)
                    if location is not None:
                        # print(f"CREDIT NOTE {cn.DocNumber} ASSIGNED TO {c.Suffix} LOCATION ID {location.Id}")
                        cn.DepartmentRef = location.to_ref()
                        cn.save(qb=self.client)
                    else:
                        print(f"[ERROR] Location IS NONE FOR {c.Suffix} OF {c}")
                else:
                    print(f"[ERROR] c.Suffix IS NONE FOR {c}")
        print(f"set_location_in_creditnotes ended.")

    def _load_excel_file(self, path):
        """Load Excel file"""
        wb = load_workbook(path)
        return wb.active

    def _create_excel_file(self, title="Undefined"):
        """Load Excel file"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = title
        return wb

    def _get_sales_term(self, id):
        """
        Get information about Sales Terms

        Name
        DayOfMonthDue
        DueDays
        """
        return Term.get(id, qb=self.client)

    def _get_customer_from_vat(self, vat_number):
        """Get customer from VAT ID"""
        customer = None
        for c in self.customers:
            if c.AlternatePhone is not None:
                if vat_number in c.AlternatePhone.FreeFormNumber:
                    # print(f"TROVATO QUALCOSA per {vat_number}")
                    customer = c
        return customer

    def _get_customer_terms(self, customer):
        """Get Customer terms of payment"""
        if customer.SalesTermRef is not None and customer.SalesTermRef.value is not None:
            return Term.get(customer.SalesTermRef.value, qb=self.client)
        return None

    def _format_description(self, data, expenses=False):
        """Format description concatenating columns"""
        if expenses:
            return f"Rimborso spese incasso"
        return f"{data[7]} Modello {data[11]} Matricola {data[10]} CTR {data[5]}/{data[4]}"

    def _get_due_date(self, terms, ref):
        # print(f"{term.Name}\t{term.DayOfMonthDue}\t{term.DueNextMonthDays}\t{term.DueDays}")
        if terms is None:
            print(f">>>>> TERMINI NULLI REF: {ref}")
            return "XXX-XX-XXXX"
        if terms.DueNextMonthDays is not None:
            months = terms.DueNextMonthDays / 30
            return f"=EOMONTH({ref}, {months})"
        return f"={ref}+{terms.DueDays}"

    # def update_invoices_terms_and_due_date(self):
    #     """Read and update an Excel file to be used to import into QB"""
    #     path = '/Users/ale/Documents/git/python-quickbooks/data/invoices.xlsx'
    #     ws = self._load_excel_file(path)

    #     for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
    #         if row[0] is not None:
    #             t = self._get_customer_terms(vat_number=row[0])


    # def list_terms(self):
    #     terms = Term.filter(Active=True, qb=self.client)
    #     for term in terms:
    #         print(f"{term.Name}\t{term.DayOfMonthDue}\t{term.DueNextMonthDays}\t{term.DueDays}")

    # def customer_check(self):
    #     for c in self.customers:
    #         if c.AlternatePhone is None:
    #             print(f"[{c.DisplayName}] PARTITA IVA VUOTA {c.PrimaryTaxIdentifier}")
    #         if c.PrimaryTaxIdentifier.startswith("XXX"):
    #             print(f"[{c.DisplayName}] PARTITA IVA ERRATA {c.PrimaryTaxIdentifier}")

    def import_invoices(self, input='data/invoices_origin.xlsx', output='data/output.xlsx'):
        """
        Import invoices into Quickbooks

        1. Open input file (INVOICES.xlsx) and set active sheet
        2. Create output file (TO_BE_IMPORTED.xlsx)
        2. For each row
            2.1 Select customer from VAT number (or CF if empty)
            2.2 Get customer Terms of Payment (ToP)
            2.3 Evaluate formula for due date based on ToP
            2.3 Write row into TO_BE_IMPORTED.xlsx
        """
        # 1
        ws = self._load_excel_file(input)
        # 2
        wb_out = self._create_excel_file(title="Fatture da importare")
        ws_out = wb_out.active
        self.headings(ws_out)
        row_no = 2

        for row in ws.iter_rows(min_row=2, min_col=1, max_col=12, values_only=True):
            if row[0] is not None:
                # 2.1
                vat_id = f"IT{row[1]}"
                if row[1] == '':
                    vat_id = f"CF{row[2]}"

                customer = self._get_customer_from_vat(vat_number=vat_id)

                if customer is None:
                    print(f"[E - {row[0]}] CLIENTE NON TROVATO VAT ID: {vat_id}")
                else:
                    # 2.2
                    terms = self._get_customer_terms(customer)
                    if terms is None:
                        print(f"[E - {customer}] TERMINI DI PAGAMENTO NON TROVATI")
                    self._output(ws=ws_out, data=row, terms=terms, row_no=row_no)
                    if row[3] == 'S':
                        # add row for expenses
                        row_no += 1
                        self._output(
                            ws=ws_out,
                            data=row,
                            terms=terms,
                            row_no=row_no,
                            expenses=True
                        )
            row_no += 1

        wb_out.save(output)

    def _output(self, ws, data, terms, row_no, expenses=False):
        import datetime
        # # Customer
        ws.cell(column=1, row=row_no, value=data[0])
        # # Partita Iva
        # vat = f"IT{data[1]}"
        # if data[1] == '':
        #     vat = f"CF{data[2]}"
        # ws.cell(column=2, row=row_no, value=vat)
        # # Invoice no
        # ws.cell(column=3, row=row_no, value='')


        # Invoice date
        today = datetime.date.today().strftime('%d/%m/%Y')
        ws.cell(column=2, row=row_no, value=today)
        # Due date
        ref = ws.cell(column=2, row=row_no).coordinate
        ws.cell(column=3, row=row_no, value=self._get_due_date(terms, ref))
        # Terms
        ws.cell(column=4, row=row_no, value=f"{terms}")


        # # Item product/service
        # ws.cell(column=7, row=row_no, value=data[6])
        # # Item description
        # ws.cell(column=8, row=row_no, value=self._format_description(data, expenses=expenses))
        # # Amount
        # ws.cell(column=9, row=row_no, value=data[10])
        # ws.cell(column=10, row=row_no, value='')

        # if expenses:
        #     ws.cell(column=7, row=row_no, value="SP")
        #     ws.cell(column=9, row=row_no, value="4,50")

    def headings(self, ws):
        """
        Customer
        Partita Iva
        Invoice no                  EMPTY
        Invoice date                EMPTY
        Due date                    =
        Terms
        Item product/service
        Item description            EVALUATED
        Amount
        Item Tax Code
        """
        row = 1
        ws.cell(column=1, row=row, value='Customer')
        # ws.cell(column=2, row=row, value='Partita Iva')
        # ws.cell(column=3, row=row, value='Invoice no')
        ws.cell(column=2, row=row, value='Invoice date')
        ws.cell(column=3, row=row, value='Due date')
        ws.cell(column=4, row=row, value='Terms')
        # ws.cell(column=7, row=row, value='Item product/service')
        # ws.cell(column=8, row=row, value='Item description')
        # ws.cell(column=9, row=row, value='Amount')
        # ws.cell(column=10, row=row, value='Item Tax Code')

    def list_credit_notes(self):
        creditnotes = CreditMemo.all(order_by="DocNumber DESC", qb=self.client)
        for cn in creditnotes:
            print(f"Nota di Credito n {cn.DepartmentRef}")

    def set_location_in_accounting(self):
        self._set_location_in_invoices()
        self._set_location_in_creditnotes()


if __name__ == '__main__':

    if len(sys.argv) < 2:
        usage()
        sys.exit(2)
    main(sys.argv[1:])

    pyqb = PythonQuickBooks()

    if _action == 'invoices':
        input_file = _arguments["-i"]
        output_file = _arguments["-o"]

        pyqb.import_invoices(input=input_file, output=output_file)

    elif _action == 'locations':
        pyqb.set_location_in_accounting()

    else:
        usage()
        sys.exit(2)